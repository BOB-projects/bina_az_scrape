#!/usr/bin/env python3
"""
Bina.az Rent Property Scraper
Scrapes all rental properties from bina.az using asyncio and aiohttp for optimal performance.
"""

import asyncio
import aiohttp
import json
import csv
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from urllib.parse import urlencode
import sys
from pathlib import Path
import time
import re

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('rent_scraper.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Import shared utilities
from scraper_utils import extract_category_from_html

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not installed. XLSX export will not be available. Install with: pip install openpyxl")


def extract_category_from_html(html: str) -> Optional[str]:
    """Extract category from detail page HTML - ONLY looks for 'Kateqoriya' field
    
    Returns specific property categories like:
    - 'Yeni tikili' (New building)
    - 'Köhnə tikili' (Old building)
    - 'Apartament'
    - etc.
    
    Returns None if Kateqoriya field is not found (no generic fallback).
    """
    import re
    
    # ONLY extract from the specific Kateqoriya product property
    # This captures categories like "Yeni tikili", "İşlənmiş", "Apartament", etc.
    kateqoriya_match = re.search(
        r'<label class="product-properties__i-name">Kateqoriya</label>\s*<span class="product-properties__i-value">([^<]+)</span>',
        html
    )
    if kateqoriya_match:
        category = kateqoriya_match.group(1).strip()
        if category and len(category) < 100:
            return category
    
    return None  # Return None if Kateqoriya not found (no fallback)


class BinaRentScraper:
    """Asynchronous scraper for bina.az rental properties"""

    BASE_URL = "https://bina.az/graphql"
    OPERATION_NAME = "SearchItems"
    SHA256_HASH = "872e9c694c34b6674514d48e9dcf1b46241d3d79f365ddf20d138f18e74554c5"

    # Pagination settings
    ITEMS_PER_PAGE = 16  # Maximum allowed by API complexity limit
    MAX_CONCURRENT_REQUESTS = 5  # Limit concurrent requests
    RETRY_ATTEMPTS = 3
    RETRY_DELAY = 2  # seconds

    # Data safety settings
    CHECKPOINT_INTERVAL = 50  # Save checkpoint every N pages
    INCREMENTAL_SAVE_INTERVAL = 100  # Save data every N pages

    def __init__(self, output_dir: str = "data/rent", resume: bool = True):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.session: Optional[aiohttp.ClientSession] = None
        self.all_items: List[Dict] = []
        self.seen_ids: set = set()  # Track IDs for deduplication
        self.semaphore = asyncio.Semaphore(self.MAX_CONCURRENT_REQUESTS)
        self.checkpoint_file = self.output_dir / "checkpoint.json"
        self.resume_enabled = resume

        # Performance tracking
        self.start_time: Optional[float] = None
        self.resume_time: Optional[float] = None
        self.page_times: List[float] = []  # Track time per page for ETA
        self.last_progress_log: float = 0
        self.progress_log_interval: int = 10  # Log detailed progress every N pages

    async def __aenter__(self):
        """Async context manager entry"""
        connector = aiohttp.TCPConnector(ssl=False)
        self.session = aiohttp.ClientSession(
            connector=connector,
            headers={
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': '*/*',
                'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8,ru;q=0.7,az;q=0.6',
                'Content-Type': 'application/json',
                'Referer': 'https://bina.az/kiraye',
                'Origin': 'https://bina.az',
            },
            timeout=aiohttp.ClientTimeout(total=30)
        )
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Async context manager exit"""
        if self.session:
            await self.session.close()

    def build_url(self, cursor: Optional[str] = None) -> str:
        """Build GraphQL API URL with parameters"""
        variables = {
            "first": self.ITEMS_PER_PAGE,
            "filter": {"leased": True},
            "sort": "BUMPED_AT_DESC"
        }

        if cursor:
            variables["cursor"] = cursor

        params = {
            "operationName": self.OPERATION_NAME,
            "variables": json.dumps(variables, separators=(',', ':')),
            "extensions": json.dumps({
                "persistedQuery": {
                    "version": 1,
                    "sha256Hash": self.SHA256_HASH
                }
            }, separators=(',', ':'))
        }

        return f"{self.BASE_URL}?{urlencode(params)}"

    def extract_item_data(self, node: Dict) -> Dict[str, Any]:
        """Extract and flatten item data from node"""
        item = {
            # Basic information
            'id': node.get('id'),
            'area_value': node.get('area', {}).get('value') if node.get('area') else None,
            'area_units': node.get('area', {}).get('units') if node.get('area') else None,
            'leased': node.get('leased'),
            'floor': node.get('floor'),
            'floors': node.get('floors'),
            'rooms': node.get('rooms'),

            # Location information
            'city_id': node.get('city', {}).get('id') if node.get('city') else None,
            'city_name': node.get('city', {}).get('name') if node.get('city') else None,
            'location_id': node.get('location', {}).get('id') if node.get('location') else None,
            'location_name': node.get('location', {}).get('name') if node.get('location') else None,
            'location_full_name': node.get('location', {}).get('fullName') if node.get('location') else None,

            # Price information
            'price_value': node.get('price', {}).get('value') if node.get('price') else None,
            'price_currency': node.get('price', {}).get('currency') if node.get('price') else None,

            # Company/Agent information
            'company_id': node.get('company', {}).get('id') if node.get('company') else None,
            'company_name': node.get('company', {}).get('name') if node.get('company') else None,
            'company_target_type': node.get('company', {}).get('targetType') if node.get('company') else None,

            # Property features
            'has_mortgage': node.get('hasMortgage'),
            'has_bill_of_sale': node.get('hasBillOfSale'),
            'has_repair': node.get('hasRepair'),
            'paid_daily': node.get('paidDaily'),
            'is_business': node.get('isBusiness'),

            # Promotion status
            'vipped': node.get('vipped'),
            'featured': node.get('featured'),

            # Metadata
            'updated_at': node.get('updatedAt'),
            'path': node.get('path'),
            'photos_count': node.get('photosCount'),

            # Photos URLs
            'photos': json.dumps([
                {
                    'thumbnail': photo.get('thumbnail'),
                    'f460x345': photo.get('f460x345'),
                    'large': photo.get('large')
                }
                for photo in node.get('photos', [])
            ]),

            # Full URL
            'url': f"https://bina.az{node.get('path')}" if node.get('path') else None,
            
            # Category (will be filled later)
            'category': None,

            # Scraping metadata
            'scraped_at': datetime.now().isoformat()
        }

        return item

    def validate_item(self, item: Dict) -> bool:
        """Validate that item has required fields"""
        required_fields = ['id', 'scraped_at']
        return all(item.get(field) is not None for field in required_fields)

    def load_checkpoint(self) -> Optional[Dict]:
        """Load checkpoint from file"""
        if not self.resume_enabled or not self.checkpoint_file.exists():
            return None

        try:
            with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                checkpoint = json.load(f)
            logger.info(f"Loaded checkpoint: {checkpoint['items_count']} items, cursor: {checkpoint.get('cursor', 'None')[:50]}...")
            return checkpoint
        except Exception as e:
            logger.warning(f"Failed to load checkpoint: {e}")
            return None

    def save_checkpoint(self, cursor: Optional[str], page_num: int, total_count: int):
        """Save checkpoint to file"""
        try:
            checkpoint = {
                'cursor': cursor,
                'page_num': page_num,
                'items_count': len(self.all_items),
                'total_count': total_count,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(checkpoint, f, ensure_ascii=False, indent=2)
            logger.debug(f"Checkpoint saved: page {page_num}, {len(self.all_items)} items")
        except Exception as e:
            logger.error(f"Failed to save checkpoint: {e}")

    def save_incremental(self, page_num: int):
        """Save incremental backup of data"""
        try:
            filename = f"backup_rent_page{page_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = self.output_dir / filename
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self.all_items, f, ensure_ascii=False, indent=2)
            logger.info(f"Incremental backup saved: {filepath} ({len(self.all_items)} items)")
        except Exception as e:
            logger.error(f"Failed to save incremental backup: {e}")

    def load_from_checkpoint(self):
        """Load previously scraped data from checkpoint"""
        checkpoint = self.load_checkpoint()
        if not checkpoint:
            return None, 0, 0

        # Look for the most recent backup file
        backup_files = sorted(self.output_dir.glob("backup_rent_page*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
        if backup_files:
            try:
                with open(backup_files[0], 'r', encoding='utf-8') as f:
                    self.all_items = json.load(f)
                # Rebuild seen_ids set
                self.seen_ids = {item['id'] for item in self.all_items if item.get('id')}
                logger.info(f"Loaded {len(self.all_items)} items from backup: {backup_files[0].name}")
            except Exception as e:
                logger.error(f"Failed to load backup file: {e}")

        return checkpoint.get('cursor'), checkpoint.get('page_num', 0), checkpoint.get('total_count', 0)

    def format_time_detailed(self, seconds: float) -> str:
        """Format seconds into detailed human-readable time"""
        if seconds < 0:
            return "calculating..."
        hours, remainder = divmod(int(seconds), 3600)
        minutes, secs = divmod(remainder, 60)
        if hours > 0:
            return f"{hours}h {minutes}m {secs}s"
        elif minutes > 0:
            return f"{minutes}m {secs}s"
        else:
            return f"{secs}s"

    def calculate_eta(self, current_page: int, total_pages: int) -> tuple[float, float]:
        """Calculate ETA and speed metrics"""
        if not self.page_times or current_page == 0:
            return 0, 0
        recent_times = self.page_times[-20:] if len(self.page_times) > 20 else self.page_times
        avg_time_per_page = sum(recent_times) / len(recent_times)
        pages_remaining = total_pages - current_page
        estimated_seconds = pages_remaining * avg_time_per_page
        return estimated_seconds, avg_time_per_page

    def log_progress(self, page_num: int, total_count: int, items_added: int, items_skipped: int, force: bool = False):
        """Log detailed progress with performance metrics"""
        current_time = time.time()
        total_pages = (total_count + self.ITEMS_PER_PAGE - 1) // self.ITEMS_PER_PAGE
        progress_pct = (page_num / total_pages * 100) if total_pages > 0 else 0

        logger.info(f"Page {page_num}/{total_pages}: +{items_added} items, {items_skipped} skipped | Total: {len(self.all_items)} / {total_count} ({progress_pct:.1f}%)")

        if force or (page_num % self.progress_log_interval == 0 and page_num > 0):
            elapsed = current_time - self.start_time
            eta_seconds, avg_time_per_page = self.calculate_eta(page_num, total_pages)
            
            pages_per_minute = (page_num / elapsed) * 60 if elapsed > 0 else 0
            items_per_minute = (len(self.all_items) / elapsed) * 60 if elapsed > 0 else 0

            eta_str = (datetime.now() + timedelta(seconds=eta_seconds)).strftime('%H:%M:%S') if eta_seconds > 0 else "calculating..."

            logger.info("=" * 80)
            logger.info(f"PROGRESS REPORT - Page {page_num}/{total_pages}")
            logger.info("-" * 80)
            logger.info(f"Progress:        [{progress_pct:5.1f}%] {page_num}/{total_pages} pages")
            logger.info(f"Items scraped:   {len(self.all_items):,} / {total_count:,}")
            logger.info(f"Elapsed time:    {self.format_time_detailed(elapsed)}")
            logger.info(f"Time remaining:  {self.format_time_detailed(eta_seconds)}")
            logger.info(f"ETA:             {eta_str}")
            logger.info(f"Speed:           {pages_per_minute:.1f} pages/min | {items_per_minute:.1f} items/min")
            logger.info("=" * 80)

    async def fetch_page(self, cursor: Optional[str] = None, attempt: int = 1) -> Optional[Dict]:
        """Fetch a single page of results with retry logic"""
        async with self.semaphore:
            url = self.build_url(cursor)
            try:
                async with self.session.get(url) as response:
                    if response.status == 200:
                        return await response.json()
                    elif attempt < self.RETRY_ATTEMPTS:
                        await asyncio.sleep(self.RETRY_DELAY * attempt)
                        return await self.fetch_page(cursor, attempt + 1)
                    return None
            except Exception as e:
                if attempt < self.RETRY_ATTEMPTS:
                    await asyncio.sleep(self.RETRY_DELAY * attempt)
                    return await self.fetch_page(cursor, attempt + 1)
                return None

    async def fetch_item_category(self, item_path: str) -> Optional[str]:
        """Fetch and extract category from item detail page"""
        if not item_path:
            return None
        
        try:
            detail_url = f"https://bina.az{item_path}"
            async with self.session.get(detail_url) as response:
                if response.status == 200:
                    html = await response.text()
                    return extract_category_from_html(html)
        except Exception as e:
            logger.debug(f"Failed to fetch category for {item_path}: {e}")
        
        return None

    async def scrape_all(self) -> List[Dict]:
        """Scrape all items with pagination and crash recovery"""
        self.start_time = time.time()
        cursor, page_num, total_count = self.load_from_checkpoint()

        if cursor:
            logger.info(f"Resuming from checkpoint: page {page_num}, {len(self.all_items)} items already scraped")
        else:
            logger.info("Starting fresh rent scrape...")
            page_num = 0
            total_count = 0

        consecutive_failures = 0
        max_consecutive_failures = 5

        try:
            while True:
                page_start_time = time.time()
                page_num += 1
                data = await self.fetch_page(cursor)

                if not data or 'data' not in data or data['data'] is None:
                    consecutive_failures += 1
                    if consecutive_failures >= max_consecutive_failures:
                        break
                    await asyncio.sleep(5 * consecutive_failures)
                    page_num -= 1
                    continue

                consecutive_failures = 0
                items_connection = data['data'].get('itemsConnection')
                if not items_connection:
                    break

                if page_num == 1 or total_count == 0:
                    total_count = items_connection.get('totalCount', 0)
                    logger.info(f"Total items to scrape: {total_count}")

                edges = items_connection.get('edges', [])
                if not edges:
                    break

                items_added = 0
                items_skipped = 0
                for edge in edges:
                    node = edge.get('node')
                    if node:
                        item_data = self.extract_item_data(node)
                        if not self.validate_item(item_data) or item_data['id'] in self.seen_ids:
                            items_skipped += 1
                            continue
                        
                        # Fetch category from detail page
                        item_path = node.get('path')
                        if item_path:
                            category = await self.fetch_item_category(item_path)
                            item_data['category'] = category
                        
                        self.all_items.append(item_data)
                        self.seen_ids.add(item_data['id'])
                        items_added += 1

                self.page_times.append(time.time() - page_start_time)
                self.log_progress(page_num, total_count, items_added, items_skipped)

                if page_num % self.CHECKPOINT_INTERVAL == 0:
                    self.save_checkpoint(cursor, page_num, total_count)
                if page_num % self.INCREMENTAL_SAVE_INTERVAL == 0:
                    self.save_incremental(page_num)

                page_info = items_connection.get('pageInfo', {})
                if not page_info.get('hasNextPage', False):
                    break
                cursor = page_info.get('endCursor')
                if not cursor:
                    break
                await asyncio.sleep(0.5)

        except Exception as e:
            logger.error(f"Error during scraping: {e}")
            raise
        finally:
            self.save_checkpoint(cursor, page_num, total_count)

        return self.all_items

    def save_to_json(self, filename: str = None):
        if filename is None:
            filename = f"bina_rent_{datetime.now().strftime('%Y%m')}.json"
        filepath = self.output_dir / filename
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.all_items, f, ensure_ascii=False, indent=2)
        logger.info(f"Data saved to JSON: {filepath}")
        return filepath

    def save_to_csv(self, filename: str = None):
        if not self.all_items: return None
        if filename is None:
            filename = f"bina_rent_{datetime.now().strftime('%Y%m')}.csv"
        filepath = self.output_dir / filename
        fieldnames = list(self.all_items[0].keys())
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.all_items)
        logger.info(f"Data saved to CSV: {filepath}")
        return filepath

    def save_to_xlsx(self, filename: str = None):
        """Save data to Excel (XLSX) file"""
        if not XLSX_AVAILABLE:
            logger.warning("openpyxl not installed. Cannot save to XLSX format.")
            return None

        if not self.all_items:
            logger.warning("No data to save")
            return None

        if filename is None:
            filename = f"bina_rent_{datetime.now().strftime('%Y%m')}.xlsx"

        filepath = self.output_dir / filename

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Rentals"

        # Get headers from first item
        headers = list(self.all_items[0].keys())

        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_num, item in enumerate(self.all_items, 2):
            for col_num, header in enumerate(headers, 1):
                value = item.get(header)
                # Convert boolean to string for better Excel compatibility
                if isinstance(value, bool):
                    value = "Yes" if value else "No"
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(filepath)
        logger.info(f"Data saved to XLSX: {filepath}")
        return filepath

async def main():
    logger.info("=" * 80)
    logger.info("Bina.az Rent Property Scraper")
    logger.info("=" * 80)

    async with BinaRentScraper() as scraper:
        items = await scraper.scrape_all()
        if items:
            scraper.save_to_json()
            scraper.save_to_csv()
            scraper.save_to_xlsx()
            logger.info(f"Scraping completed. Total items: {len(items)}")
        else:
            logger.error("No items scraped!")

if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}")
        sys.exit(1)
