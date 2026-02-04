#!/usr/bin/env python3
"""
Bina.az Sale Property Scraper
Scrapes all sale properties from bina.az using asyncio and aiohttp for optimal performance.
"""

import asyncio
import aiohttp
import json
import csv
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Optional, Any
from urllib.parse import urlencode
import sys
from pathlib import Path
import time
import re

# Configure logging first
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('scraper.log', encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Import shared utilities
from scraper_utils import extract_category_from_html

# Try to import openpyxl for Excel support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    logger.warning("openpyxl not installed. XLSX export will not be available. Install with: pip install openpyxl")


def extract_category_from_html(html: str) -> Optional[str]:
    """Extract category from detail page HTML - looks for 'Kateqoriya' field"""
    import re
    
    # Look for the two main categories first: Yeni tikili or Köhnə tikili
    # Pattern 1: Direct text search (most reliable)
    if 'Yeni tikili' in html:
        return 'Yeni tikili'
    elif 'Köhnə tikili' in html:
        return 'Köhnə tikili'
    
    # Pattern 2: Look for "Kateqoriya" label followed by the value (capture any category)
    kateqoriya_match = re.search(r'Kateqoriya["\']?\s*(?:</?\w+[^>]*>)*\s*([^<>\n]+?)(?:</|$)', html, re.IGNORECASE)
    if kateqoriya_match:
        category = kateqoriya_match.group(1).strip()
        if category and len(category) < 100:  # Sanity check
            return category
    
    # Pattern 3: Look in data attributes or JSON
    data_match = re.search(r'"category"\s*:\s*"([^"]+)"', html)
    if data_match:
        return data_match.group(1)
    
    return None  # Return None if not found


class BinaScraper:
    """Asynchronous scraper for bina.az properties"""

    BASE_URL = "https://bina.az/graphql"
    OPERATION_NAME = "SearchItems"
    SHA256_HASH = "872e9c694c34b6674514d48e9dcf1b46241d3d79f365ddf20d138f18e74554c5"

    # Pagination settings
    ITEMS_PER_PAGE = 16  # Maximum allowed by API complexity limit
    MAX_CONCURRENT_REQUESTS = 5  # Limit concurrent requests
    RETRY_ATTEMPTS = 3
    RETRY_DELAY = 2  # seconds

    # Data safety settings
    CHECKPOINT_INTERVAL = 50  # Save checkpoint every N pages
    INCREMENTAL_SAVE_INTERVAL = 100  # Save data every N pages

    def __init__(self, output_dir: str = "data", resume: bool = True):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.session: Optional[aiohttp.ClientSession] = None
        self.all_items: List[Dict] = []
        self.seen_ids: set = set()  # Track IDs for deduplication
        self.semaphore = asyncio.Semaphore(self.MAX_CONCURRENT_REQUESTS)
        self.checkpoint_file = self.output_dir / "checkpoint.json"
        self.resume_enabled = resume

        # Performance tracking
        self.start_time: Optional[float] = None
        self.resume_time: Optional[float] = None
        self.page_times: List[float] = []  # Track time per page for ETA
        self.last_progress_log: float = 0
        self.progress_log_interval: int = 10  # Log detailed progress every N pages

    async def __aenter__(self):
        """Async context manager entry"""
        connector = aiohttp.TCPConnector(ssl=False)
        self.session = aiohttp.ClientSession(
            connector=connector,
            headers={
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
                'Accept': '*/*',
                'Accept-Language': 'en-GB,en-US;q=0.9,en;q=0.8,ru;q=0.7,az;q=0.6',
                'Content-Type': 'application/json',
                'Referer': 'https://bina.az/alqi-satqi',
                'Origin': 'https://bina.az',
            },
            timeout=aiohttp.ClientTimeout(total=30)
        )
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Async context manager exit"""
        if self.session:
            await self.session.close()

    def build_url(self, cursor: Optional[str] = None) -> str:
        """Build GraphQL API URL with parameters"""
        variables = {
            "first": self.ITEMS_PER_PAGE,
            "filter": {"leased": False},
            "sort": "BUMPED_AT_DESC"
        }

        if cursor:
            variables["cursor"] = cursor

        params = {
            "operationName": self.OPERATION_NAME,
            "variables": json.dumps(variables, separators=(',', ':')),
            "extensions": json.dumps({
                "persistedQuery": {
                    "version": 1,
                    "sha256Hash": self.SHA256_HASH
                }
            }, separators=(',', ':'))
        }

        return f"{self.BASE_URL}?{urlencode(params)}"

    def extract_item_data(self, node: Dict) -> Dict[str, Any]:
        """Extract and flatten item data from node"""
        item = {
            # Basic information
            'id': node.get('id'),
            'area_value': node.get('area', {}).get('value') if node.get('area') else None,
            'area_units': node.get('area', {}).get('units') if node.get('area') else None,
            'leased': node.get('leased'),
            'floor': node.get('floor'),
            'floors': node.get('floors'),
            'rooms': node.get('rooms'),

            # Location information
            'city_id': node.get('city', {}).get('id') if node.get('city') else None,
            'city_name': node.get('city', {}).get('name') if node.get('city') else None,
            'location_id': node.get('location', {}).get('id') if node.get('location') else None,
            'location_name': node.get('location', {}).get('name') if node.get('location') else None,
            'location_full_name': node.get('location', {}).get('fullName') if node.get('location') else None,

            # Price information
            'price_value': node.get('price', {}).get('value') if node.get('price') else None,
            'price_currency': node.get('price', {}).get('currency') if node.get('price') else None,

            # Company/Agent information
            'company_id': node.get('company', {}).get('id') if node.get('company') else None,
            'company_name': node.get('company', {}).get('name') if node.get('company') else None,
            'company_target_type': node.get('company', {}).get('targetType') if node.get('company') else None,

            # Property features
            'has_mortgage': node.get('hasMortgage'),
            'has_bill_of_sale': node.get('hasBillOfSale'),
            'has_repair': node.get('hasRepair'),
            'paid_daily': node.get('paidDaily'),
            'is_business': node.get('isBusiness'),

            # Promotion status
            'vipped': node.get('vipped'),
            'featured': node.get('featured'),

            # Metadata
            'updated_at': node.get('updatedAt'),
            'path': node.get('path'),
            'photos_count': node.get('photosCount'),

            # Photos URLs
            'photos': json.dumps([
                {
                    'thumbnail': photo.get('thumbnail'),
                    'f460x345': photo.get('f460x345'),
                    'large': photo.get('large')
                }
                for photo in node.get('photos', [])
            ]),

            # Full URL
            'url': f"https://bina.az{node.get('path')}" if node.get('path') else None,
            
            # Category (will be filled later)
            'category': None,

            # Scraping metadata
            'scraped_at': datetime.now().isoformat()
        }

        return item

    def validate_item(self, item: Dict) -> bool:
        """Validate that item has required fields"""
        required_fields = ['id', 'scraped_at']
        return all(item.get(field) is not None for field in required_fields)

    def load_checkpoint(self) -> Optional[Dict]:
        """Load checkpoint from file"""
        if not self.resume_enabled or not self.checkpoint_file.exists():
            return None

        try:
            with open(self.checkpoint_file, 'r', encoding='utf-8') as f:
                checkpoint = json.load(f)
            logger.info(f"Loaded checkpoint: {checkpoint['items_count']} items, cursor: {checkpoint.get('cursor', 'None')[:50]}...")
            return checkpoint
        except Exception as e:
            logger.warning(f"Failed to load checkpoint: {e}")
            return None

    def save_checkpoint(self, cursor: Optional[str], page_num: int, total_count: int):
        """Save checkpoint to file"""
        try:
            checkpoint = {
                'cursor': cursor,
                'page_num': page_num,
                'items_count': len(self.all_items),
                'total_count': total_count,
                'timestamp': datetime.now().isoformat()
            }
            with open(self.checkpoint_file, 'w', encoding='utf-8') as f:
                json.dump(checkpoint, f, ensure_ascii=False, indent=2)
            logger.debug(f"Checkpoint saved: page {page_num}, {len(self.all_items)} items")
        except Exception as e:
            logger.error(f"Failed to save checkpoint: {e}")

    def save_incremental(self, page_num: int):
        """Save incremental backup of data"""
        try:
            filename = f"backup_page{page_num}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            filepath = self.output_dir / filename
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(self.all_items, f, ensure_ascii=False, indent=2)
            logger.info(f"Incremental backup saved: {filepath} ({len(self.all_items)} items)")
        except Exception as e:
            logger.error(f"Failed to save incremental backup: {e}")

    def load_from_checkpoint(self):
        """Load previously scraped data from checkpoint"""
        checkpoint = self.load_checkpoint()
        if not checkpoint:
            return None, 0, 0

        # Look for the most recent backup file (sort by modification time)
        backup_files = sorted(self.output_dir.glob("backup_page*.json"), key=lambda x: x.stat().st_mtime, reverse=True)
        if backup_files:
            try:
                with open(backup_files[0], 'r', encoding='utf-8') as f:
                    self.all_items = json.load(f)
                # Rebuild seen_ids set
                self.seen_ids = {item['id'] for item in self.all_items if item.get('id')}
                logger.info(f"Loaded {len(self.all_items)} items from backup: {backup_files[0].name}")
            except Exception as e:
                logger.error(f"Failed to load backup file: {e}")

        return checkpoint.get('cursor'), checkpoint.get('page_num', 0), checkpoint.get('total_count', 0)

    def format_time(self, seconds: float) -> str:
        """Format seconds into human-readable time"""
        if seconds < 60:
            return f"{seconds:.0f}s"
        elif seconds < 3600:
            minutes = seconds / 60
            return f"{minutes:.1f}m"
        else:
            hours = seconds / 3600
            return f"{hours:.1f}h"

    def format_time_detailed(self, seconds: float) -> str:
        """Format seconds into detailed human-readable time"""
        if seconds < 0:
            return "calculating..."

        hours, remainder = divmod(int(seconds), 3600)
        minutes, secs = divmod(remainder, 60)

        if hours > 0:
            return f"{hours}h {minutes}m {secs}s"
        elif minutes > 0:
            return f"{minutes}m {secs}s"
        else:
            return f"{secs}s"

    def calculate_eta(self, current_page: int, total_pages: int) -> tuple[float, float]:
        """Calculate ETA and speed metrics"""
        if not self.page_times or current_page == 0:
            return 0, 0

        # Use recent pages for more accurate estimation (last 20 pages)
        recent_times = self.page_times[-20:] if len(self.page_times) > 20 else self.page_times
        avg_time_per_page = sum(recent_times) / len(recent_times)

        pages_remaining = total_pages - current_page
        estimated_seconds = pages_remaining * avg_time_per_page

        return estimated_seconds, avg_time_per_page

    def log_progress(self, page_num: int, total_count: int, items_added: int, items_skipped: int, force: bool = False):
        """Log detailed progress with performance metrics"""
        current_time = time.time()

        # Calculate total pages
        total_pages = (total_count + self.ITEMS_PER_PAGE - 1) // self.ITEMS_PER_PAGE
        pages_remaining = total_pages - page_num
        progress_pct = (page_num / total_pages * 100) if total_pages > 0 else 0

        # Log standard progress every page
        logger.info(f"Page {page_num}/{total_pages}: +{items_added} items, {items_skipped} skipped | Total: {len(self.all_items)} / {total_count} ({progress_pct:.1f}%)")

        # Log detailed progress every N pages or when forced
        if force or (page_num % self.progress_log_interval == 0 and page_num > 0):
            elapsed = current_time - self.start_time

            # Calculate performance metrics
            eta_seconds, avg_time_per_page = self.calculate_eta(page_num, total_pages)

            if page_num > 0:
                pages_per_minute = (page_num / elapsed) * 60
                items_per_minute = (len(self.all_items) / elapsed) * 60
            else:
                pages_per_minute = 0
                items_per_minute = 0

            # Calculate ETA timestamp
            if eta_seconds > 0:
                eta_time = datetime.now() + timedelta(seconds=eta_seconds)
                eta_str = eta_time.strftime('%H:%M:%S')
            else:
                eta_str = "calculating..."

            # Log detailed progress
            logger.info("=" * 80)
            logger.info(f"PROGRESS REPORT - Page {page_num}/{total_pages}")
            logger.info("-" * 80)
            logger.info(f"Progress:        [{progress_pct:5.1f}%] {page_num}/{total_pages} pages")
            logger.info(f"Items scraped:   {len(self.all_items):,} / {total_count:,}")
            logger.info(f"Pages remaining: {pages_remaining:,}")
            logger.info(f"")
            logger.info(f"TIME METRICS")
            logger.info(f"Elapsed time:    {self.format_time_detailed(elapsed)}")
            logger.info(f"Time remaining:  {self.format_time_detailed(eta_seconds)}")
            logger.info(f"ETA:             {eta_str}")
            logger.info(f"")
            logger.info(f"PERFORMANCE")
            logger.info(f"Speed:           {pages_per_minute:.1f} pages/min")
            logger.info(f"                 {items_per_minute:.1f} items/min")
            logger.info(f"Avg per page:    {self.format_time(avg_time_per_page)}")
            logger.info("=" * 80)

    async def fetch_page(self, cursor: Optional[str] = None, attempt: int = 1) -> Optional[Dict]:
        """Fetch a single page of results with retry logic"""
        async with self.semaphore:
            url = self.build_url(cursor)

            try:
                async with self.session.get(url) as response:
                    if response.status == 200:
                        data = await response.json()
                        return data
                    else:
                        logger.warning(f"Request failed with status {response.status}")

                        if attempt < self.RETRY_ATTEMPTS:
                            await asyncio.sleep(self.RETRY_DELAY * attempt)
                            return await self.fetch_page(cursor, attempt + 1)

                        return None

            except asyncio.TimeoutError:
                logger.error(f"Timeout error for cursor: {cursor}")
                if attempt < self.RETRY_ATTEMPTS:
                    await asyncio.sleep(self.RETRY_DELAY * attempt)
                    return await self.fetch_page(cursor, attempt + 1)
                return None

            except Exception as e:
                logger.error(f"Error fetching page: {e}", exc_info=True)
                if attempt < self.RETRY_ATTEMPTS:
                    await asyncio.sleep(self.RETRY_DELAY * attempt)
                    return await self.fetch_page(cursor, attempt + 1)
                return None

    async def fetch_item_category(self, item_path: str) -> Optional[str]:
        """Fetch and extract category from item detail page"""
        if not item_path:
            return None
        
        try:
            detail_url = f"https://bina.az{item_path}"
            async with self.session.get(detail_url) as response:
                if response.status == 200:
                    html = await response.text()
                    return extract_category_from_html(html)
        except Exception as e:
            logger.debug(f"Failed to fetch category for {item_path}: {e}")
        
        return None

    async def scrape_all(self) -> List[Dict]:
        """Scrape all items with pagination and crash recovery"""
        # Initialize start time
        self.start_time = time.time()

        # Try to resume from checkpoint
        cursor, page_num, total_count = self.load_from_checkpoint()

        if cursor:
            logger.info(f"Resuming from checkpoint: page {page_num}, {len(self.all_items)} items already scraped")
            self.resume_time = time.time()
        else:
            logger.info("Starting fresh scrape...")
            page_num = 0
            total_count = 0

        consecutive_failures = 0
        max_consecutive_failures = 5

        try:
            while True:
                page_start_time = time.time()
                page_num += 1

                data = await self.fetch_page(cursor)

                if not data or 'data' not in data or data['data'] is None:
                    consecutive_failures += 1
                    logger.error(f"Failed to fetch data or invalid response (failure {consecutive_failures}/{max_consecutive_failures})")

                    if consecutive_failures >= max_consecutive_failures:
                        logger.error("Max consecutive failures reached. Stopping scraper.")
                        self.save_checkpoint(cursor, page_num - 1, total_count)
                        self.save_incremental(page_num - 1)
                        break

                    # Save checkpoint before retrying
                    self.save_checkpoint(cursor, page_num - 1, total_count)
                    page_num -= 1  # Don't increment page number on failure
                    # Retry after delay
                    logger.info(f"Retrying after {5 * consecutive_failures} seconds...")
                    await asyncio.sleep(5 * consecutive_failures)
                    continue

                # Reset failure counter on success
                consecutive_failures = 0

                items_connection = data['data'].get('itemsConnection')
                if not items_connection:
                    logger.error("No itemsConnection in response")
                    self.save_checkpoint(cursor, page_num, total_count)
                    break

                # Get total count on first page
                if page_num == 1 or total_count == 0:
                    total_count = items_connection.get('totalCount', 0)
                    logger.info(f"Total items to scrape: {total_count}")

                # Extract items with validation and deduplication
                edges = items_connection.get('edges', [])
                if not edges:
                    logger.info("No more items found")
                    break

                items_added = 0
                items_skipped = 0

                for edge in edges:
                    node = edge.get('node')
                    if node:
                        item_data = self.extract_item_data(node)

                        # Validate item
                        if not self.validate_item(item_data):
                            logger.warning(f"Invalid item skipped: {item_data.get('id', 'unknown')}")
                            items_skipped += 1
                            continue

                        # Check for duplicates
                        item_id = item_data['id']
                        if item_id in self.seen_ids:
                            logger.debug(f"Duplicate item skipped: {item_id}")
                            items_skipped += 1
                            continue
                        
                        # Fetch category from detail page
                        item_path = node.get('path')
                        if item_path:
                            category = await self.fetch_item_category(item_path)
                            item_data['category'] = category

                        # Add item
                        self.all_items.append(item_data)
                        self.seen_ids.add(item_id)
                        items_added += 1

                # Track page time for ETA calculation
                page_time = time.time() - page_start_time
                self.page_times.append(page_time)

                # Log progress with performance metrics
                self.log_progress(page_num, total_count, items_added, items_skipped)

                # Save checkpoint periodically
                if page_num % self.CHECKPOINT_INTERVAL == 0:
                    self.save_checkpoint(cursor, page_num, total_count)

                # Save incremental backup periodically
                if page_num % self.INCREMENTAL_SAVE_INTERVAL == 0:
                    self.save_incremental(page_num)

                # Check if there are more pages
                page_info = items_connection.get('pageInfo', {})
                has_next_page = page_info.get('hasNextPage', False)

                if not has_next_page:
                    logger.info("Reached last page")
                    break

                # Get cursor for next page
                cursor = page_info.get('endCursor')
                if not cursor:
                    logger.warning("No cursor for next page")
                    break

                # Small delay to be respectful
                await asyncio.sleep(0.5)

        except KeyboardInterrupt:
            logger.warning("Scraping interrupted by user - saving checkpoint...")
            self.save_checkpoint(cursor, page_num, total_count)
            self.save_incremental(page_num)
            raise
        except Exception as e:
            logger.error(f"Error during scraping: {e}", exc_info=True)
            logger.info("Saving checkpoint before exiting...")
            self.save_checkpoint(cursor, page_num, total_count)
            self.save_incremental(page_num)
            raise
        finally:
            # Final checkpoint and cleanup
            self.save_checkpoint(cursor, page_num, total_count)

        # Final performance summary
        total_time = time.time() - self.start_time
        total_pages = (total_count + self.ITEMS_PER_PAGE - 1) // self.ITEMS_PER_PAGE

        logger.info("\n" + "=" * 80)
        logger.info("SCRAPING COMPLETED!")
        logger.info("=" * 80)
        logger.info(f"Total items scraped:    {len(self.all_items):,}")
        logger.info(f"Unique items:           {len(self.seen_ids):,}")
        logger.info(f"Duplicates removed:     {len(self.all_items) - len(self.seen_ids):,}")
        logger.info(f"Total pages processed:  {page_num:,} / {total_pages:,}")
        logger.info(f"")
        logger.info(f"FINAL TIME METRICS")
        logger.info(f"Total time:             {self.format_time_detailed(total_time)}")
        logger.info(f"Average per page:       {self.format_time(total_time / page_num if page_num > 0 else 0)}")
        logger.info(f"Average per item:       {self.format_time(total_time / len(self.all_items) if self.all_items else 0)}")
        logger.info(f"")
        logger.info(f"FINAL PERFORMANCE")
        logger.info(f"Pages per minute:       {(page_num / total_time * 60):.1f}")
        logger.info(f"Items per minute:       {(len(self.all_items) / total_time * 60):.1f}")
        logger.info(f"Items per second:       {(len(self.all_items) / total_time):.2f}")
        logger.info("=" * 80 + "\n")

        return self.all_items

    def save_to_json(self, filename: str = None):
        """Save data to JSON file"""
        if filename is None:
            filename = f"bina_sale_{datetime.now().strftime('%Y%m')}.json"

        filepath = self.output_dir / filename

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(self.all_items, f, ensure_ascii=False, indent=2)

        logger.info(f"Data saved to JSON: {filepath}")
        return filepath

    def save_to_csv(self, filename: str = None):
        """Save data to CSV file"""
        if not self.all_items:
            logger.warning("No data to save")
            return

        if filename is None:
            filename = f"bina_sale_{datetime.now().strftime('%Y%m')}.csv"

        filepath = self.output_dir / filename

        # Get all keys from first item
        fieldnames = list(self.all_items[0].keys())

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(self.all_items)

        logger.info(f"Data saved to CSV: {filepath}")
        return filepath

    def save_to_xlsx(self, filename: str = None):
        """Save data to Excel (XLSX) file"""
        if not XLSX_AVAILABLE:
            logger.warning("openpyxl not installed. Cannot save to XLSX format.")
            return None

        if not self.all_items:
            logger.warning("No data to save")
            return None

        if filename is None:
            filename = f"bina_sale_{datetime.now().strftime('%Y%m')}.xlsx"

        filepath = self.output_dir / filename

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Properties"

        # Get headers from first item
        headers = list(self.all_items[0].keys())

        # Style for header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        for row_num, item in enumerate(self.all_items, 2):
            for col_num, header in enumerate(headers, 1):
                value = item.get(header)
                # Convert boolean to string for better Excel compatibility
                if isinstance(value, bool):
                    value = "Yes" if value else "No"
                ws.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save workbook
        wb.save(filepath)
        logger.info(f"Data saved to XLSX: {filepath}")
        return filepath

    def get_statistics(self) -> Dict[str, Any]:
        """Get statistics about scraped data"""
        if not self.all_items:
            return {}

        total = len(self.all_items)

        # Calculate statistics
        stats = {
            'total_items': total,
            'with_photos': sum(1 for item in self.all_items if item.get('photos_count', 0) > 0),
            'with_mortgage': sum(1 for item in self.all_items if item.get('has_mortgage')),
            'with_repair': sum(1 for item in self.all_items if item.get('has_repair')),
            'vipped': sum(1 for item in self.all_items if item.get('vipped')),
            'featured': sum(1 for item in self.all_items if item.get('featured')),
            'business': sum(1 for item in self.all_items if item.get('is_business')),
        }

        # Price statistics
        prices = [item['price_value'] for item in self.all_items if item.get('price_value')]
        if prices:
            stats['avg_price'] = sum(prices) / len(prices)
            stats['min_price'] = min(prices)
            stats['max_price'] = max(prices)

        # Area statistics
        areas = [item['area_value'] for item in self.all_items if item.get('area_value')]
        if areas:
            stats['avg_area'] = sum(areas) / len(areas)
            stats['min_area'] = min(areas)
            stats['max_area'] = max(areas)

        # City distribution
        cities = {}
        for item in self.all_items:
            city = item.get('city_name')
            if city:
                cities[city] = cities.get(city, 0) + 1
        stats['cities'] = cities

        # Room distribution
        rooms = {}
        for item in self.all_items:
            room = item.get('rooms')
            if room:
                rooms[str(room)] = rooms.get(str(room), 0) + 1
        stats['rooms_distribution'] = rooms

        return stats

    def cleanup_backups(self, keep_last: int = 3):
        """Clean up old backup files, keeping only the last N"""
        try:
            backup_files = sorted(self.output_dir.glob("backup_page*.json"))
            if len(backup_files) > keep_last:
                for backup_file in backup_files[:-keep_last]:
                    backup_file.unlink()
                    logger.info(f"Removed old backup: {backup_file.name}")
        except Exception as e:
            logger.warning(f"Failed to cleanup backups: {e}")

    def get_data_integrity_report(self) -> Dict[str, Any]:
        """Generate data integrity report"""
        if not self.all_items:
            return {}

        report = {
            'total_items': len(self.all_items),
            'unique_items': len(self.seen_ids),
            'duplicates_removed': len(self.all_items) - len(self.seen_ids),
            'items_with_missing_price': sum(1 for item in self.all_items if not item.get('price_value')),
            'items_with_missing_location': sum(1 for item in self.all_items if not item.get('location_name')),
            'items_with_missing_area': sum(1 for item in self.all_items if not item.get('area_value')),
            'items_with_photos': sum(1 for item in self.all_items if item.get('photos_count', 0) > 0),
            'data_completeness': {
                'id': sum(1 for item in self.all_items if item.get('id')) / len(self.all_items) * 100,
                'price': sum(1 for item in self.all_items if item.get('price_value')) / len(self.all_items) * 100,
                'location': sum(1 for item in self.all_items if item.get('location_name')) / len(self.all_items) * 100,
                'area': sum(1 for item in self.all_items if item.get('area_value')) / len(self.all_items) * 100,
            }
        }
        return report


async def main():
    """Main function to run the scraper"""
    logger.info("=" * 80)
    logger.info("Bina.az Sale Property Scraper")
    logger.info("=" * 80)

    async with BinaScraper() as scraper:
        # Scrape all data
        items = await scraper.scrape_all()

        if items:
            # Cleanup old backups
            scraper.cleanup_backups(keep_last=3)

            # Save to JSON, CSV, and XLSX formats
            json_file = scraper.save_to_json()
            csv_file = scraper.save_to_csv()
            xlsx_file = scraper.save_to_xlsx()

            # Print data integrity report
            integrity = scraper.get_data_integrity_report()
            logger.info("\n" + "=" * 80)
            logger.info("DATA INTEGRITY REPORT")
            logger.info("=" * 80)
            logger.info(f"Total items scraped: {integrity.get('total_items', 0)}")
            logger.info(f"Unique items: {integrity.get('unique_items', 0)}")
            logger.info(f"Duplicates removed: {integrity.get('duplicates_removed', 0)}")
            logger.info(f"\nData Completeness:")
            for field, completeness in integrity.get('data_completeness', {}).items():
                logger.info(f"  {field}: {completeness:.1f}%")
            logger.info(f"\nData Quality:")
            logger.info(f"  Items with photos: {integrity.get('items_with_photos', 0)}")
            logger.info(f"  Missing price: {integrity.get('items_with_missing_price', 0)}")
            logger.info(f"  Missing location: {integrity.get('items_with_missing_location', 0)}")
            logger.info(f"  Missing area: {integrity.get('items_with_missing_area', 0)}")

            # Print statistics
            stats = scraper.get_statistics()
            logger.info("\n" + "=" * 80)
            logger.info("SCRAPING STATISTICS")
            logger.info("=" * 80)
            logger.info(f"Total unique items: {stats.get('total_items', 0)}")
            logger.info(f"Items with photos: {stats.get('with_photos', 0)}")
            logger.info(f"Items with mortgage: {stats.get('with_mortgage', 0)}")
            logger.info(f"Items with repair: {stats.get('with_repair', 0)}")
            logger.info(f"VIP items: {stats.get('vipped', 0)}")
            logger.info(f"Featured items: {stats.get('featured', 0)}")
            logger.info(f"Business listings: {stats.get('business', 0)}")

            if 'avg_price' in stats:
                logger.info(f"\nPrice range: {stats['min_price']:,.0f} - {stats['max_price']:,.0f} AZN")
                logger.info(f"Average price: {stats['avg_price']:,.0f} AZN")

            if 'avg_area' in stats:
                logger.info(f"\nArea range: {stats['min_area']:.1f} - {stats['max_area']:.1f} m2")
                logger.info(f"Average area: {stats['avg_area']:.1f} m2")

            logger.info("\nTop cities:")
            for city, count in sorted(stats.get('cities', {}).items(),
                                     key=lambda x: x[1], reverse=True)[:10]:
                logger.info(f"  {city}: {count} listings")

            logger.info("\nRoom distribution:")
            for rooms, count in sorted(stats.get('rooms_distribution', {}).items()):
                logger.info(f"  {rooms} rooms: {count} listings")

            logger.info("=" * 80)
            logger.info(f"Files saved:")
            logger.info(f"  JSON: {json_file}")
            logger.info(f"  CSV: {csv_file}")
            if xlsx_file:
                logger.info(f"  XLSX: {xlsx_file}")
            logger.info("=" * 80)
        else:
            logger.error("No items scraped!")


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("\nScraping interrupted by user")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        sys.exit(1)
